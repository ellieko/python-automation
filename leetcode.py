from cgitb import text
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By

wb = Workbook()
ws = wb.active
ws.title = "top100"

browser = webdriver.Chrome()
browser.get("https://leetcode.com/problem-list/top-interview-questions")

ws.append(["Title", "Link", "Difficulty"])
for i in range(1, 51):
    ws.cell(row=i+1, column=1, value=browser.find_element(By.XPATH, f"//*[@id=\"__next\"]/div/div/div/div[1]/div[2]/div[2]/div/div/div[2]/div[{i}]/div[2]/div/div/div/div/a").text)
    ws.cell(row=i+1, column=2).hyperlink = browser.find_element(By.LINK_TEXT, ws.cell(row=i+1, column=1).value).get_attribute("href")
    ws.cell(row=i+1, column=3, value=browser.find_element(By.XPATH, f"//*[@id=\"__next\"]/div/div/div/div[1]/div[2]/div[2]/div/div/div[2]/div[{i}]/div[5]/span").text)
wb.save("leetcode.xlsx")
wb.close()